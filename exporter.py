import io
import logging
from typing import Optional

import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

def _get_col_decimals(ws, min_row):
    """Определяет максимальное количество знаков после запятой для каждого столбца."""
    col_decimals = {}
    for col in ws.columns:
        col_idx = col[0].column
        max_d = 0
        for cell in col:
            if cell.row < min_row:
                continue
            if isinstance(cell.value, float):
                s = str(cell.value)
                if '.' in s:
                    decimals = len(s.split('.')[1])
                    if decimals == 1 and s.split('.')[1] == '0':
                        pass
                    elif decimals > max_d:
                        max_d = decimals
        col_decimals[col_idx] = max_d
    return col_decimals

def _format_details_sheet(ws):
    """Применяет форматирование к листу Детализация: фильтры, ширину колонок, стили."""
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    ws.freeze_panes = 'A2'
    
    date_cols = set()
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        if isinstance(cell.value, str) and 'дата' in cell.value.lower():
            date_cols.add(col_idx)
            
    if ws.max_row > 1 and ws.max_column > 0:
        coord = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        ws.auto_filter.ref = coord

    col_decimals = _get_col_decimals(ws, min_row=2)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = zebra_fill if row_idx % 2 == 0 else white_fill
        for cell in row:
            cell.fill = fill
            cell.border = thin_border
            
            # Превращаем специальные строки в кликабельные гиперссылки
            if isinstance(cell.value, str) and "|||" in cell.value:
                url, text = cell.value.split("|||", 1)
                cell.value = text
                if url:
                    cell.hyperlink = url
                    cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")
            
            # Центральное выравнивание для чисел и дат
            if cell.column in date_cols or isinstance(cell.value, (int, float)):
                 cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                 cell.alignment = Alignment(vertical="center")
                 
            if cell.column in date_cols and cell.value:
                 cell.number_format = 'DD.MM.YYYY HH:MM'
            elif isinstance(cell.value, float):
                 max_d = col_decimals.get(cell.column, 0)
                 cell.number_format = '#,##0.' + '0' * max_d if max_d > 0 else '#,##0'
            elif isinstance(cell.value, int):
                 cell.number_format = '#,##0'

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    val_str = str(cell.value)
                    for line in val_str.split('\n'):
                        if len(line) > max_length:
                            max_length = len(line)
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 3, 50) # ограничиваем макс ширину

def _format_summary_sheet(ws, date_min: str, date_max: str, title: str):
    """Премиальная верстка для сводных листов (Итого)."""
    # Пишем заголовок
    ws.insert_rows(1, 2) # Отодвигаем таблицу на 2 строки вниз
    ws["A1"] = f"{title}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="003366")
    ws["A2"] = f"Отчет за период: с {date_min} по {date_max}"
    ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="555555")
    
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F8FF", end_color="F2F8FF", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
                         
    max_col = ws.max_column
    max_row = ws.max_row
    
    # Шапка таблицы (находится на 3 строке)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    col_decimals = _get_col_decimals(ws, min_row=4)
        
    # Данные таблицы
    for row in range(4, max_row + 1):
        is_total_row = False
        first_cell_val = str(ws.cell(row=row, column=1).value)
        if "ИТОГО" in first_cell_val.upper():
            is_total_row = True
            
        bg_fill = zebra_fill if row % 2 == 0 else white_fill
        if is_total_row:
            bg_fill = PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid")
            
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = bg_fill
            if is_total_row:
                cell.font = Font(bold=True)
            
            if col > 1 or isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")
                
            if isinstance(cell.value, float):
                max_d = col_decimals.get(col, 0)
                cell.number_format = '#,##0.' + '0' * max_d if max_d > 0 else '#,##0'
            elif isinstance(cell.value, int):
                cell.number_format = '#,##0'

    # Ширина колонок
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.row < 3: # игнорируем заголовки листа
                    continue
                if cell.value:
                    val_str = str(cell.value)
                    for line in val_str.split('\n'):
                        if len(line) > max_length:
                            max_length = len(line)
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 4

def export_report_to_excel(df_details: pd.DataFrame, df_summary_corp: pd.DataFrame, df_summary_cars: pd.DataFrame, date_min: str, date_max: str, output_path: Optional[str] = None) -> Optional[bytes]:
    """Формирует Excel-отчет из датафреймов с детальной и сводной статистикой.
    
    Args:
        df_details: DataFrame со всеми рассчитанными рейсами.
        df_summary_corp: DataFrame со сводными метриками по общим итогам.
        df_summary_cars: DataFrame со сводными метриками по машинам.
        date_min: Строка с минимальной датой для заголовка.
        date_max: Строка с максимальной датой для заголовка.
        output_path: Опциональный путь к файлу. Если None, метод вернет байты файла.
        
    Returns:
        Байты Excel файла (если output_path is None) или None (файл сохранен).
    """
    if output_path:
        logger.info(f"Начало экспорта отчета в файл {output_path}...")
        writer_obj = output_path
    else:
        logger.info("Формирование отчета в памяти...")
        writer_obj = io.BytesIO()
        
    with pd.ExcelWriter(writer_obj, engine='openpyxl') as writer:
        if not df_summary_corp.empty:
            df_summary_corp.to_excel(writer, sheet_name="Итого по контрагентам", index=False)
        if not df_summary_cars.empty:
            df_summary_cars.to_excel(writer, sheet_name="Итого по машинам", index=False)
            
        if not df_details.empty:
            df_details.to_excel(writer, sheet_name="Детализация", index=False)
            
        wb = writer.book
        
        if "Итого по контрагентам" in wb.sheetnames:
            _format_summary_sheet(wb["Итого по контрагентам"], date_min, date_max, title="Сводный отчет: Контрагенты")
            
        if "Итого по машинам" in wb.sheetnames:
            _format_summary_sheet(wb["Итого по машинам"], date_min, date_max, title="Сводный отчет: Машины (детализация)")
            
        if "Детализация" in wb.sheetnames:
            _format_details_sheet(wb["Детализация"])
            
    if not output_path:
        writer_obj.seek(0)
        return writer_obj.read()
